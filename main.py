from openpyxl.utils import get_column_letter as getColLetters
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table
from openpyxl import load_workbook


def getRange(ws, rg=None):
    """
    Get the usable row and column range of a worksheet.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
        rg (dict, optional): A dictionary with keys 'r0', 'r1', 'c0', 'c1'.
            - 'r0': Minimum row
            - 'r1': Maximum row
            - 'c0': Minimum column
            - 'c1': Maximum column
            If any value is None or missing, the worksheet's min/max values are used.

    Returns:
        dict: A dictionary with normalized values for the range:
              { 'r0': int, 'r1': int, 'c0': int, 'c1': int }
    """
    r0 = rg.get('r0') if rg.get(
        'r0') is not None and rg is not None else ws.min_row
    r1 = rg.get('r1') if rg.get(
        'r1') is not None and rg is not None else ws.max_row
    c0 = rg.get('c0') if rg.get(
        'c0') is not None and rg is not None else ws.min_col
    c1 = rg.get('c1') if rg.get(
        'c1') is not None and rg is not None else ws.max_col

    return {
        'r0': r0, 'r1': r1, 'c0': c0, 'c1': c1
    }


def AltA(ws):
    """
    Get the full data range of a worksheet as an Excel reference string.

    Args:
        ws (Worksheet): The openpyxl worksheet object.

    Returns:
        str: Excel-style range reference (e.g., "A1:D20")
    """
    r1 = ws.max_row
    c1 = ws.max_column

    return f"A1:{getColLetters(c1)}{r1}"


def AltHOI(ws):
    """
    Auto-adjust column widths based on cell contents.

    Args:
        ws (Worksheet): The openpyxl worksheet object.

    Notes:
        - Width is based on the longest string length in each column.
        - Adds +3 padding to avoid truncated values.
    """
    colWidths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                # Calculate the maximum length of each col
                colWidths[cell.column_letter] = max(
                    (colWidths.get(cell.column_letter, 0)),
                    len(str(cell.value))
                )
    # Add a little extra width
    for col, width in colWidths.items():
        ws.column_dimensions[col].width = width + 3


def AltHK(ws, rg=None):
    """
    Apply numeric formatting with thousands separators and parentheses for negatives.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
        rg (dict, optional): Range dictionary with keys 'r0', 'r1', 'c0', 'c1'.
            If not provided, the full worksheet is used.

    Format Applied:
        "#,##0.00;(#,##0.00);-"
    """
    # if list of range (rg) is missing, use max ranges
    rg = getRange(ws)

    for col in ws.iter_cols(
            min_row=rg['r0'], max_row=rg['r1'],
            min_col=rg['c0'], max_col=rg['c1']):
        for cell in col:
            cell.number_format = '#,##0.00;(#,##0.00);-'


def AltHN_Cus(ws, temp, rg=None):
    """
    Apply a custom numeric format to a worksheet range.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
        temp (str): Excel number format string (e.g. "#,##0.00").
        rg (dict, optional): Range dictionary with keys 'r0', 'r1', 'c0', 'c1'.
            If not provided, the full worksheet is used.
    """
    # if list of range (rg) is missing, use max ranges
    rg = getRange(ws)

    for col in ws.iter_cols(
            min_row=rg['r0'], max_row=rg['r1'],
            min_col=rg['c0'], max_col=rg['c1']):
        for cell in col:
            cell.number_format = f'{temp}'


def AltHNS(ws, rg=None):
    """
    Apply a standard date format ("MM/DD/YYYY") to all date/datetime cells.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
        rg (dict, optional): Range dictionary with keys 'r0', 'r1', 'c0', 'c1'.
            If not provided, the full worksheet is used.
    """
    # if list of range (rg) is missing, use max ranges
    rg = getRange(ws)

    for col in ws.iter_cols(
            min_row=rg['r0'], max_row=rg['r1'],
            min_col=rg['c0'], max_col=rg['c1']):
        for cell in col:
            # Check for cell value being a date or datetime
            if cell.is_date:
                # Apply date format MM/DD/YYYY
                cell.number_format = 'mm/dd/yyyy'


def AltHNS_Cus(ws, temp, rg=None):
    """
    Apply a custom date format to all date/datetime cells in a range.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
        temp (str): Excel date format string (e.g. "DD-MMM-YYYY").
        rg (dict, optional): Range dictionary with keys 'r0', 'r1', 'c0', 'c1'.
            If not provided, the full worksheet is used.
    """
    # if list of range (rg) is missing, use max ranges
    rg = getRange(ws)

    for col in ws.iter_cols(
            min_row=rg['r0'], max_row=rg['r1'],
            min_col=rg['c0'], max_col=rg['c1']):
        for cell in col:
            # Check for cell value being a date or datetime
            if cell.is_date:
                # Apply date format MM/DD/YYYY
                cell.number_format = f'{temp}'


def tableFormatWB(wb):
    """
    Apply table formatting to all worksheets in a workbook.

    Actions Performed:
        - Freeze first row (A2).
        - Left-align header row.
        - Hide gridlines.
        - Convert full sheet data into a table object.
        - Auto-adjust column widths.

    Args:
        wb (Workbook): The openpyxl workbook object.
    """
    for sh in wb.sheetnames:
        ws = wb[sh]
        ws.freeze_panes = 'A2'
        for col in ws.iter_cols(max_row=1, max_col=ws.max_column):
            for cell in col:
                cell.alignment = Alignment(horizontal='left')
        ws.sheet_view.showGridLines = False
        wsData = AltA(ws)
        ws.add_table(Table(displayName=f'{sh}', ref=wsData))
        ws = AltHOI(ws)


def tableFormatWS(ws):
    """
    Apply table formatting to a single worksheet.

    Actions Performed:
        - Freeze first row (A2).
        - Left-align header row.
        - Hide gridlines.
        - Convert full sheet data into a table object.
        - Auto-adjust column widths.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
    """
    ws.freeze_panes = 'A2'
    for col in ws.iter_cols(max_row=1, max_col=ws.max_column):
        for cell in col:
            cell.alignment = Alignment(horizontal='left')
    ws.sheet_view.showGridLines = False
    wsData = AltA(ws)
    ws.add_table(Table(displayName=f'{ws.title}', ref=wsData))
    ws = AltHOI(ws)


def formatWS(wb, shz=None):
    """
    Apply table formatting to specific sheets in a workbook.

    Args:
        wb (Workbook): The openpyxl workbook object.
        shz (list[str], optional): List of worksheet names to format.
            If None, applies to all sheets.
    """
    # if list of sheets not provided, pick all sheets in wb
    if shz is None:
        shz = wb.sheetnames

    # loop through all sheets in the given list
    for sh in shz:
        ws = wb[sh]
        tableFormatWS(ws)


def formatWB(wbNin, wbNout=None, ty=None, shz=None):
    """
    High-level workbook formatting controller.

    Args:
        wbNin (str): Input workbook filename.
        wbNout (str, optional): Output workbook filename.
            If None, overwrites the input workbook.
        ty (str, optional): Type of formatting:
            - None: Apply full table formatting to all sheets.
            - 'ws': Apply table formatting to specific sheets.
            - 'sp': Apply specific numeric/date formatting to hardcoded sheets.
        shz (list[str], optional): List of sheet names (used only if ty='ws').

    Notes:
        - Uses load_workbook to open the Excel file.
        - Saves formatted workbook to wbNout.
    """
    wb = load_workbook(wbNin)

    # Format all sheets in wb with sheetname loop
    if ty is None:
        tableFormatWB(wb)

    # Format all sheets in wb with additional control
    if ty == 'ws':
        formatWS(wb, shz)

    # Specific column based formatting
    if ty == 'sp':
        AltHK(wb['common'], {'r0': 2, 'r1': 4, 'c0': 3, 'c1': 3})
        AltHNS(wb['added'], {'r0': 2, 'r1': 3, 'c0': 2, 'c1': 2})

    # To replace the formatted wb with original
    if wbNout is None:
        wbNout = wbNin

    wb.save(wbNout)


def main():
    wbNin = 'fileInput.xlsx'
    wbNout = 'fileOutput.xlsx'

    # Variable to control formatting function choice
    ty = 'ws'

    # Creating the list of desired sheets
    shz = ['added', 'common']

    # Formate sheets accordingly
    formatWB(wbNin, wbNout)
    formatWB(wbNin, wbNout, ty, shz)


if __name__ == '__main__':
    main()
